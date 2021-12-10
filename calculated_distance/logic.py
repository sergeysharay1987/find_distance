import os
from typing import List, Tuple, Union

import numpy
import pandas
import shapely.geometry
from geopandas import GeoDataFrame
import geopandas
from geopy import Location
from openpyxl import load_workbook
from loguru import logger
from geopy.geocoders import Yandex
from shapely.geometry import Polygon, Point
from shapely.ops import nearest_points
from geopy.distance import geodesic
import matplotlib as plt

API_KEY = 'cbddbd2c-95ce-4aa1-ba5a-5d0416597c20'
ya_geocoder: Yandex = Yandex(api_key=API_KEY)  # геокодер, используемый для геокодирования адреса
number_of_mkad_s_kms: int = 108  # кол-во километров МКАД
mkad_address: str = 'Россия Москва МКАД'  # неизменяющееся часть адреса МКАД, используемая для формирования адреса
# каждого км МКАД
number_of_km: int = 0  # номер километра МКАД, начиная с нулевого
coords_mkad: list = []
url: str = f'https://geocode-maps.yandex.ru/1.x/'
blueprint: str = 'calculated_distance'
excel_file: str = 'coords_of_mkad_s_kms.xlsx'
shape_file = 'dataframe.shp'
list_mkad_s_km: List[float] = []  # список, для хранения координат каждого километра МКАД
dir_to_excel: str = os.getcwd()  # путь до .xlsx файла
dir_to_shape_file: str = os.getcwd()


def make_list_of_addresses_for_request():
    """Возвращает строку, содержащую адреса всех километров МКАД, разделённых запятой"""
    many_addresses = ''
    splitter = ','
    for number_of_km in range(0, number_of_mkad_s_kms + 1):
        many_addresses += f'{mkad_address} {number_of_km} kilometr vneshnayay storona{splitter}'
    return many_addresses


# print(make_list_of_addresses_for_request())

def check_content():
    """Проверяет содержит ли файл координаты всех километров МКАД"""
    wb = load_workbook(excel_file)  # открываем .xlsx файл
    sheet = wb.active  # получаем текущий активный лист
    for col in sheet.iter_cols(min_col=1, max_col=2, min_row=2, max_row=number_of_mkad_s_kms + 2,
                               values_only=True):
        if 'POINT (' not in col and '0123456789.' not in col[7:-1]:
            return False
    return True


def create_geodataframe() -> GeoDataFrame:
    """Возвращает GeoDataFrame, содержащий координаты всех километров МКАД"""
    list_of_addresses = make_list_of_addresses_for_request().split(',')
    # Создаём GeoDataFrame, содержащий координаты всех километров МКАД
    coords_of_mkad_s_kms = geopandas.tools.geocode(list_of_addresses, provider='yandex',
                                                   api_key=API_KEY)
    return coords_of_mkad_s_kms


def get_coords(ex_file: str) -> List[float]:
    """Возвращает список, содержащий вложенные списки с координатами каждого километра МКАД"""

    wb = load_workbook(f'{dir_to_excel}/{ex_file}')
    sheet_1 = wb.active
    for row in sheet_1.iter_rows(min_row=2, max_col=3, max_row=number_of_mkad_s_kms + 2, values_only=True):
        # получаем координаты точек для построения полигона МКАД
        coords = row[1].split('(')[-1].replace(')', '').split()
        coords = list(map(float, coords))
        # меняем местами координаты долготу с широтой, первая координата - широта, вторая - долгота
        # для получения правильного результата с помощью функции geopy.distance.geodesic
        coords[0], coords[1] = coords[1], coords[0]
        coords_mkad.append(coords)
    return coords_mkad


# def get_coords_polygon() -> List[float]:
#     """Возвращает список координат всех километров МКАД для построения полигона"""
#     # проверяем есть ли в текущей директории
#     if shape_file in os.listdir(dir_to_excel):
#         # открываем файл,
#         return get_coords(excel_file)
#     # проверяем есть ли excel файл в текущей директории
#     elif excel_file not in os.listdir('.') or excel_file in os.listdir('.') and check_content():
#
#         # создаём .xlsx файл, если его не было в текущей директории и сохраняем в него координаты всех километров МКАД
#         create_geodataframe().to_excel(
#             excel_writer=f'{dir_to_excel}/{excel_file}')
#         return get_coords(excel_file)


def get_coords_polygon() -> List[float]:
    """Возвращает список координат всех километров МКАД для построения полигона"""
    # проверяем есть ли в текущей директории
    if shape_file in os.listdir(dir_to_shape_file):
        # открываем файл,

        return create_geodataframe().to_file(shape_file)
    # проверяем есть ли файл .shp в текущей директории
    elif shape_file not in os.listdir('.') or shape_file in os.listdir('.'):
        # создаём .shp файл, если его не было в текущей директории и сохраняем в него координаты всех километров МКАД
        return create_geodataframe().to_file(shape_file)


def geocode_address(address: str) -> Tuple[float, float]:
    """Возвращает кортеж, содержащий широту и долготу"""
    coords: Location = ya_geocoder.geocode(address)
    return coords.latitude, coords.longitude


def write_in_log(address: str, distance: float):
    path = f'{os.getcwd()}/{blueprint}/info.log'
    logger.add(path, format='{time} {message}', level='INFO')
    logger.info(f'Расстояние от МКАД до {address} равно {distance} км')


def find_distance(coords_of_address) -> Union[int, float]:
    """Возвращает расстояние в километрах от МКАД до адреса, введённого в поле формы 'адрес'"""
    # получаем координаты всех километров МКАД из .xlsx файла
    coords = get_coords_polygon()
    coords_of_address = Point(coords_of_address)
    # создаём полигон из полученных координат
    poly_mkad = Polygon(coords)
    #    g = geopandas.read_file(excel_file)
    # g.plot()
    # plt.show()
    if poly_mkad.contains(coords_of_address):
        return 0
    # находим точку на полигоне, ближайшую к точке, содержащей координаты адреса <coords_of_address>,
    # для которого требуется найти расстояние
    else:
        p1, nearest_pt = nearest_points(coords_of_address, poly_mkad)  # ищем точку (nearest_pt) на МКАД,
        # расположенную ближе всего к
        # точке, с координатами адреса, расстояние до которого требуется найти
        distance = geodesic((nearest_pt.x, nearest_pt.y), (coords_of_address.x, coords_of_address.y))
        return round(distance.km, 1)


# file = geopandas.read_file(f'{os.getcwd()}/RUS_adm/RUS_adm2.shp')
# #var = file.var()
# #print(type(file))
# #print(file.columns)
# #print(file.get('NAME_1', 'Moskva'))
# #print(file.head())
# # for name in file['NAME_1'].values:
# #     if name == 'Moskva':
# #         print(name)
# #print(file['NAME_1'].values)
# #print(file['NAME_2'].values)
# #print(var)
# file.plot()
# coords_Moscow = geocode_address('Moscow')
# find_distance(coords_Moscow)
gdf = geopandas.read_file(dir_to_shape_file + '/' + shape_file)
# gdf = gdf[]
# gdf.plot()
# gdf.drop()
# print(gdf['geometry'][0])
# for pt in gdf.address:
#     #if not isinstance(pt, Point):
#
#     print(pt)
# gdf.set_geometry(col = 0)
# print(gdf)
lat_lon_coords = []
lan_lon_each = []

for point in gdf.geometry:
    # print(point)
    if isinstance(point, Point):
        # print(point.xy)
        lan_lon_each.append(point.y)
        lan_lon_each.append(point.x)
        # print(point.y, point.x)
        lat_lon_coords.append(Point(lan_lon_each))

        lan_lon_each = []
    else:
        break
        # gdf.drop(point)
coords = numpy.array(lat_lon_coords, dtype=object)
polygon = Polygon(lat_lon_coords)

print(geocode_address('Moscow'))
print(polygon.contains(Point(geocode_address('Russia Moscow'))))
# gdf_poly = GeoDataFrame(polygon)
print(polygon)
print(coords.dtype)
print(type(coords))
gdf['geometry'] = pandas.Series(coords)
# print(gdf['geometry'])
print(gdf)
gdf.plot()

# gdf.set_geometry(lat_lon_coords)
# print(gdf)
# print(lat_lon_coords[108])
# gdf_1 = GeoDataFrame(lat_lon_coords)
# gdf_1.set_geometry(col = lat_lon_coords)
# print(gdf_1.rename())
# gdf_1.rename()
# gdf_1.set_geometry(lat_lon_coords)
# gdf[0] = gdf['geometry']
# print(gdf_1)
# gdf.set_geometry(lat_lon_coords[:108])
# print(gdf)
